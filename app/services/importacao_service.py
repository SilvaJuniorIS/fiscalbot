import csv
import json
import logging
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import settings
from app.models.contrato import Contrato, ContratoStatus
from app.models.fornecedor import Fornecedor
from app.models.secretaria import Secretaria
from app.models.user import User
from app.services.contrato_service import calcular_status

logger = logging.getLogger(__name__)

COLUNAS_OBRIGATORIAS = {
    "numero",
    "orgao",
    "objeto",
    "valor",
    "inicio",
    "termino",
    "secretaria",
    "fornecedor",
    "cnpj",
}


@dataclass
class LinhaImportacao:
    numero_linha: int
    dados: dict[str, str]


def diretorio_importacoes() -> Path:
    path = Path(settings.storage_path) / "importacoes"
    path.mkdir(parents=True, exist_ok=True)
    return path


def salvar_upload_importacao(filename: str, content: bytes) -> Path:
    suffix = Path(filename).suffix.lower()
    if suffix not in {".csv", ".xlsx"}:
        raise ValueError("Arquivo deve ser .csv ou .xlsx")
    timestamp = datetime.utcnow().strftime("%Y%m%d%H%M%S%f")
    safe_name = re.sub(r"[^A-Za-z0-9_.-]", "_", Path(filename).name)
    path = diretorio_importacoes() / f"{timestamp}_{safe_name}"
    path.write_bytes(content)
    return path


def _normalizar_header(value: Any) -> str:
    return str(value or "").strip().lower()


def _normalizar_linha(row: dict[str, Any]) -> dict[str, str]:
    return {key: str(value).strip() if value is not None else "" for key, value in row.items()}


def _ler_csv(path: Path) -> list[LinhaImportacao]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        headers = {_normalizar_header(h) for h in (reader.fieldnames or [])}
        _validar_colunas(headers)
        linhas = []
        for index, row in enumerate(reader, start=2):
            normalizada = {_normalizar_header(k): v for k, v in row.items()}
            linhas.append(LinhaImportacao(index, _normalizar_linha(normalizada)))
        return linhas


def _ler_xlsx(path: Path) -> list[LinhaImportacao]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [_normalizar_header(value) for value in next(rows, [])]
    _validar_colunas(set(headers))
    linhas = []
    for index, row in enumerate(rows, start=2):
        dados = {headers[pos]: value for pos, value in enumerate(row) if pos < len(headers)}
        linhas.append(LinhaImportacao(index, _normalizar_linha(dados)))
    workbook.close()
    return linhas


def _validar_colunas(headers: set[str]) -> None:
    faltantes = sorted(COLUNAS_OBRIGATORIAS - headers)
    if faltantes:
        raise ValueError(f"Colunas obrigatorias ausentes: {', '.join(faltantes)}")


def carregar_linhas(path: Path) -> list[LinhaImportacao]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _ler_csv(path)
    if suffix == ".xlsx":
        return _ler_xlsx(path)
    raise ValueError("Formato de arquivo nao suportado")


def _parse_date(value: str) -> date:
    value = value.strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Data invalida: {value}")


def _parse_money(value: str) -> Decimal:
    normalized = value.strip().replace("R$", "").replace(" ", "")
    if "," in normalized and "." in normalized:
        normalized = normalized.replace(".", "").replace(",", ".")
    elif "," in normalized:
        normalized = normalized.replace(",", ".")
    try:
        money = Decimal(normalized)
    except InvalidOperation as exc:
        raise ValueError(f"Valor monetario invalido: {value}") from exc
    if money < 0:
        raise ValueError("Valor nao pode ser negativo")
    return money


def _validar_cnpj(value: str) -> str:
    digits = re.sub(r"\D", "", value)
    if len(digits) != 14:
        raise ValueError(f"CNPJ invalido: {value}")
    return value.strip()


def _get_or_create_secretaria(db: Session, nome: str) -> Secretaria:
    if not nome:
        raise ValueError("Secretaria e obrigatoria")
    secretaria = db.scalar(select(Secretaria).where(Secretaria.nome == nome))
    if secretaria:
        return secretaria
    secretaria = Secretaria(nome=nome, sigla=None, is_active=True)
    db.add(secretaria)
    db.flush()
    return secretaria


def _get_or_create_fornecedor(db: Session, nome: str, cnpj: str) -> Fornecedor:
    if not nome:
        raise ValueError("Fornecedor e obrigatorio")
    fornecedor = db.scalar(select(Fornecedor).where(Fornecedor.cnpj == cnpj))
    if fornecedor:
        if not fornecedor.razao_social:
            fornecedor.razao_social = nome
        return fornecedor
    fornecedor = Fornecedor(razao_social=nome, cnpj=cnpj, is_active=True)
    db.add(fornecedor)
    db.flush()
    return fornecedor


def _get_user_by_email(db: Session, email: str) -> User | None:
    if not email:
        return None
    return db.scalar(select(User).where(User.email == email))


def _validar_escopo(secretaria: Secretaria, scoped_secretaria_ids: list[int] | None) -> None:
    if scoped_secretaria_ids is not None and secretaria.id not in scoped_secretaria_ids:
        raise ValueError("Secretaria fora do escopo do usuario")


def _linha_para_payload(
    db: Session,
    linha: LinhaImportacao,
    scoped_secretaria_ids: list[int] | None,
) -> dict[str, Any]:
    dados = linha.dados
    numero = dados.get("numero", "").strip()
    if not numero:
        raise ValueError("Numero do contrato e obrigatorio")

    inicio = _parse_date(dados.get("inicio", ""))
    termino = _parse_date(dados.get("termino", ""))
    if termino < inicio:
        raise ValueError("Termino deve ser maior ou igual ao inicio")

    secretaria = _get_or_create_secretaria(db, dados.get("secretaria", "").strip())
    _validar_escopo(secretaria, scoped_secretaria_ids)
    fornecedor = _get_or_create_fornecedor(
        db,
        dados.get("fornecedor", "").strip(),
        _validar_cnpj(dados.get("cnpj", "")),
    )
    fiscal = _get_user_by_email(db, dados.get("fiscal_email", "").strip())
    gestor = _get_user_by_email(db, dados.get("gestor_email", "").strip())
    status = dados.get("status", "").strip() or calcular_status(termino)
    if status not in {item.value for item in ContratoStatus}:
        raise ValueError(f"Status invalido: {status}")

    return {
        "numero": numero,
        "orgao": dados.get("orgao", "").strip(),
        "objeto": dados.get("objeto", "").strip(),
        "valor": _parse_money(dados.get("valor", "")),
        "inicio": inicio,
        "termino": termino,
        "secretaria_id": secretaria.id,
        "fornecedor_id": fornecedor.id,
        "fiscal_responsavel_id": fiscal.id if fiscal else None,
        "gestor_responsavel_id": gestor.id if gestor else None,
        "status": status,
        "tags": dados.get("tags", "").strip() or None,
    }


def _empty_report() -> dict[str, Any]:
    return {
        "importados": 0,
        "ignorados": 0,
        "erros": 0,
        "atualizados": 0,
        "linhas_invalidas": [],
        "detalhes": [],
    }


def _salvar_relatorios(task_id: str, report: dict[str, Any]) -> dict[str, str]:
    base = diretorio_importacoes() / task_id
    base.mkdir(parents=True, exist_ok=True)
    json_path = base / "relatorio.json"
    csv_path = base / "relatorio.csv"
    json_path.write_text(json.dumps(report, default=str, ensure_ascii=False, indent=2), encoding="utf-8")
    with csv_path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=["linha", "numero", "status", "mensagem"])
        writer.writeheader()
        writer.writerows(report["detalhes"])
    return {"json": str(json_path), "csv": str(csv_path)}


def processar_importacao_contratos(
    db: Session,
    *,
    task_id: str,
    path: str,
    modo: str,
    scoped_secretaria_ids: list[int] | None,
) -> dict[str, Any]:
    logger.info("Iniciando importacao de contratos task=%s modo=%s arquivo=%s", task_id, modo, path)
    report = _empty_report()
    try:
        linhas = carregar_linhas(Path(path))
    except ValueError as exc:
        report["erros"] = 1
        report["linhas_invalidas"].append({"linha": 1, "erro": str(exc)})
        report["detalhes"].append({"linha": 1, "numero": "", "status": "erro", "mensagem": str(exc)})
        report["arquivos"] = _salvar_relatorios(task_id, report)
        return report

    for linha in linhas:
        numero = linha.dados.get("numero", "")
        try:
            payload = _linha_para_payload(db, linha, scoped_secretaria_ids)
            contrato = db.scalar(select(Contrato).where(Contrato.numero == payload["numero"]))
            if contrato and modo == "append":
                report["ignorados"] += 1
                status = "ignorado"
                mensagem = "Contrato duplicado"
            elif contrato:
                for field, value in payload.items():
                    setattr(contrato, field, value)
                db.add(contrato)
                report["atualizados"] += 1
                status = "atualizado"
                mensagem = "Contrato atualizado"
            else:
                db.add(Contrato(**payload))
                report["importados"] += 1
                status = "importado"
                mensagem = "Contrato importado"
            report["detalhes"].append(
                {"linha": linha.numero_linha, "numero": numero, "status": status, "mensagem": mensagem}
            )
        except Exception as exc:
            logger.warning("Linha %s ignorada na importacao %s: %s", linha.numero_linha, task_id, exc)
            db.rollback()
            report["erros"] += 1
            report["linhas_invalidas"].append({"linha": linha.numero_linha, "erro": str(exc)})
            report["detalhes"].append(
                {"linha": linha.numero_linha, "numero": numero, "status": "erro", "mensagem": str(exc)}
            )
        else:
            db.commit()

    report["total_processado"] = len(linhas)
    report["arquivos"] = _salvar_relatorios(task_id, report)
    logger.info("Importacao de contratos concluida task=%s report=%s", task_id, report)
    return report
